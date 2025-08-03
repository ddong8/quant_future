"""
资金流水服务
"""
import logging
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, desc, func, text
from decimal import Decimal
import pandas as pd
import numpy as np
from ..models.account import Transaction, Account, TransactionType, TransactionStatus
from ..models.user import User
from ..models.order import Order
from ..models.position import Position
from ..core.database import get_db

logger = logging.getLogger(__name__)

class TransactionService:
    """资金流水服务"""
    
    def __init__(self, db: Session):
        self.db = db
    
    # ==================== 流水记录管理 ====================
    
    def create_transaction(self, transaction_data: Dict[str, Any]) -> Transaction:
        """创建交易流水记录"""
        try:
            transaction = Transaction(
                account_id=transaction_data['account_id'],
                transaction_id=transaction_data.get('transaction_id', self._generate_transaction_id()),
                transaction_type=TransactionType(transaction_data['transaction_type']),
                status=TransactionStatus(transaction_data.get('status', 'COMPLETED')),
                amount=Decimal(str(transaction_data['amount'])),
                currency=transaction_data.get('currency', 'USD'),
                exchange_rate=Decimal(str(transaction_data.get('exchange_rate', 1))),
                balance_before=Decimal(str(transaction_data.get('balance_before', 0))),
                balance_after=Decimal(str(transaction_data.get('balance_after', 0))),
                order_id=transaction_data.get('order_id'),
                position_id=transaction_data.get('position_id'),
                symbol=transaction_data.get('symbol'),
                description=transaction_data.get('description', ''),
                reference_id=transaction_data.get('reference_id'),
                fee_amount=Decimal(str(transaction_data.get('fee_amount', 0))),
                tax_amount=Decimal(str(transaction_data.get('tax_amount', 0))),
                metadata=transaction_data.get('metadata', {}),
                transaction_time=transaction_data.get('transaction_time', datetime.now())
            )
            
            self.db.add(transaction)
            self.db.commit()
            self.db.refresh(transaction)
            
            logger.info(f"创建交易流水成功: {transaction.transaction_id}")
            return transaction
            
        except Exception as e:
            self.db.rollback()
            logger.error(f"创建交易流水失败: {e}")
            raise
    
    def get_transaction(self, transaction_id: str) -> Optional[Transaction]:
        """获取单个交易流水"""
        return self.db.query(Transaction).filter(
            Transaction.transaction_id == transaction_id
        ).first()
    
    def get_transactions_by_account(self, account_id: int, filters: Dict[str, Any] = None,
                                  skip: int = 0, limit: int = 100) -> List[Transaction]:
        """获取账户交易流水"""
        query = self.db.query(Transaction).filter(Transaction.account_id == account_id)
        
        if filters:
            query = self._apply_transaction_filters(query, filters)
        
        return query.order_by(desc(Transaction.transaction_time)).offset(skip).limit(limit).all()
    
    def get_transactions_by_user(self, user_id: int, filters: Dict[str, Any] = None,
                               skip: int = 0, limit: int = 100) -> List[Transaction]:
        """获取用户所有账户的交易流水"""
        query = self.db.query(Transaction).join(Account).filter(Account.user_id == user_id)
        
        if filters:
            query = self._apply_transaction_filters(query, filters)
        
        return query.order_by(desc(Transaction.transaction_time)).offset(skip).limit(limit).all()
    
    def update_transaction_status(self, transaction_id: str, status: TransactionStatus,
                                metadata: Dict[str, Any] = None) -> Optional[Transaction]:
        """更新交易状态"""
        try:
            transaction = self.get_transaction(transaction_id)
            if not transaction:
                return None
            
            transaction.status = status
            if metadata:
                transaction.metadata.update(metadata)
            
            self.db.commit()
            self.db.refresh(transaction)
            
            logger.info(f"更新交易状态成功: {transaction_id} -> {status.value}")
            return transaction
            
        except Exception as e:
            self.db.rollback()
            logger.error(f"更新交易状态失败: {e}")
            raise
    
    # ==================== 流水查询和筛选 ====================
    
    def search_transactions(self, search_params: Dict[str, Any]) -> Dict[str, Any]:
        """高级搜索交易流水"""
        try:
            query = self.db.query(Transaction)
            
            # 基础筛选条件
            if 'user_id' in search_params:
                query = query.join(Account).filter(Account.user_id == search_params['user_id'])
            
            if 'account_ids' in search_params:
                query = query.filter(Transaction.account_id.in_(search_params['account_ids']))
            
            if 'transaction_types' in search_params:
                query = query.filter(Transaction.transaction_type.in_(search_params['transaction_types']))
            
            if 'status_list' in search_params:
                query = query.filter(Transaction.status.in_(search_params['status_list']))
            
            # 时间范围筛选
            if 'start_date' in search_params:
                query = query.filter(Transaction.transaction_time >= search_params['start_date'])
            
            if 'end_date' in search_params:
                query = query.filter(Transaction.transaction_time <= search_params['end_date'])
            
            # 金额范围筛选
            if 'min_amount' in search_params:
                query = query.filter(Transaction.amount >= search_params['min_amount'])
            
            if 'max_amount' in search_params:
                query = query.filter(Transaction.amount <= search_params['max_amount'])
            
            # 标的筛选
            if 'symbols' in search_params:
                query = query.filter(Transaction.symbol.in_(search_params['symbols']))
            
            # 关键词搜索
            if 'keyword' in search_params:
                keyword = f"%{search_params['keyword']}%"
                query = query.filter(
                    or_(
                        Transaction.description.ilike(keyword),
                        Transaction.reference_id.ilike(keyword),
                        Transaction.transaction_id.ilike(keyword)
                    )
                )
            
            # 获取总数
            total_count = query.count()
            
            # 排序
            sort_field = search_params.get('sort_field', 'transaction_time')
            sort_order = search_params.get('sort_order', 'desc')
            
            if sort_order == 'desc':
                query = query.order_by(desc(getattr(Transaction, sort_field)))
            else:
                query = query.order_by(getattr(Transaction, sort_field))
            
            # 分页
            skip = search_params.get('skip', 0)
            limit = search_params.get('limit', 100)
            transactions = query.offset(skip).limit(limit).all()
            
            return {
                'transactions': transactions,
                'total_count': total_count,
                'page_info': {
                    'skip': skip,
                    'limit': limit,
                    'has_more': skip + len(transactions) < total_count
                }
            }
            
        except Exception as e:
            logger.error(f"搜索交易流水失败: {e}")
            return {'transactions': [], 'total_count': 0, 'page_info': {}}
    
    def get_transaction_categories(self, user_id: int) -> Dict[str, Any]:
        """获取交易分类统计"""
        try:
            # 按交易类型分组统计
            type_stats = self.db.query(
                Transaction.transaction_type,
                func.count(Transaction.id).label('count'),
                func.sum(Transaction.amount).label('total_amount'),
                func.avg(Transaction.amount).label('avg_amount')
            ).join(Account).filter(
                Account.user_id == user_id
            ).group_by(Transaction.transaction_type).all()
            
            # 按状态分组统计
            status_stats = self.db.query(
                Transaction.status,
                func.count(Transaction.id).label('count')
            ).join(Account).filter(
                Account.user_id == user_id
            ).group_by(Transaction.status).all()
            
            # 按货币分组统计
            currency_stats = self.db.query(
                Transaction.currency,
                func.count(Transaction.id).label('count'),
                func.sum(Transaction.amount).label('total_amount')
            ).join(Account).filter(
                Account.user_id == user_id
            ).group_by(Transaction.currency).all()
            
            return {
                'by_type': [
                    {
                        'type': stat.transaction_type.value,
                        'count': stat.count,
                        'total_amount': float(stat.total_amount or 0),
                        'avg_amount': float(stat.avg_amount or 0)
                    }
                    for stat in type_stats
                ],
                'by_status': [
                    {
                        'status': stat.status.value,
                        'count': stat.count
                    }
                    for stat in status_stats
                ],
                'by_currency': [
                    {
                        'currency': stat.currency,
                        'count': stat.count,
                        'total_amount': float(stat.total_amount or 0)
                    }
                    for stat in currency_stats
                ]
            }
            
        except Exception as e:
            logger.error(f"获取交易分类统计失败: {e}")
            return {}
    
    # ==================== 统计分析 ====================
    
    def get_transaction_statistics(self, user_id: int, period: str = 'month') -> Dict[str, Any]:
        """获取交易统计分析"""
        try:
            # 确定时间范围
            end_date = datetime.now()
            if period == 'day':
                start_date = end_date - timedelta(days=1)
            elif period == 'week':
                start_date = end_date - timedelta(weeks=1)
            elif period == 'month':
                start_date = end_date - timedelta(days=30)
            elif period == 'quarter':
                start_date = end_date - timedelta(days=90)
            elif period == 'year':
                start_date = end_date - timedelta(days=365)
            else:
                start_date = end_date - timedelta(days=30)
            
            # 基础统计
            base_query = self.db.query(Transaction).join(Account).filter(
                Account.user_id == user_id,
                Transaction.transaction_time >= start_date,
                Transaction.transaction_time <= end_date,
                Transaction.status == TransactionStatus.COMPLETED
            )
            
            total_transactions = base_query.count()
            
            # 收入支出统计
            income_query = base_query.filter(Transaction.amount > 0)
            expense_query = base_query.filter(Transaction.amount < 0)
            
            total_income = income_query.with_entities(func.sum(Transaction.amount)).scalar() or Decimal('0')
            total_expense = abs(expense_query.with_entities(func.sum(Transaction.amount)).scalar() or Decimal('0'))
            net_flow = total_income - total_expense
            
            # 交易频率分析
            daily_transactions = self._get_daily_transaction_stats(user_id, start_date, end_date)
            
            # 最大单笔交易
            max_income = income_query.order_by(desc(Transaction.amount)).first()
            max_expense = expense_query.order_by(Transaction.amount).first()
            
            # 手续费统计
            total_fees = base_query.with_entities(func.sum(Transaction.fee_amount)).scalar() or Decimal('0')
            
            # 按交易类型统计
            type_breakdown = self._get_transaction_type_breakdown(user_id, start_date, end_date)
            
            return {
                'period': period,
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'summary': {
                    'total_transactions': total_transactions,
                    'total_income': float(total_income),
                    'total_expense': float(total_expense),
                    'net_flow': float(net_flow),
                    'total_fees': float(total_fees),
                    'avg_transaction_amount': float((total_income + total_expense) / total_transactions) if total_transactions > 0 else 0
                },
                'extremes': {
                    'max_income': {
                        'amount': float(max_income.amount) if max_income else 0,
                        'date': max_income.transaction_time.isoformat() if max_income else None,
                        'description': max_income.description if max_income else None
                    },
                    'max_expense': {
                        'amount': float(abs(max_expense.amount)) if max_expense else 0,
                        'date': max_expense.transaction_time.isoformat() if max_expense else None,
                        'description': max_expense.description if max_expense else None
                    }
                },
                'daily_stats': daily_transactions,
                'type_breakdown': type_breakdown
            }
            
        except Exception as e:
            logger.error(f"获取交易统计分析失败: {e}")
            return {}
    
    def get_cash_flow_analysis(self, user_id: int, period: str = 'month') -> Dict[str, Any]:
        """现金流分析"""
        try:
            # 确定时间范围
            end_date = datetime.now()
            if period == 'month':
                start_date = end_date - timedelta(days=30)
                interval = 'day'
            elif period == 'quarter':
                start_date = end_date - timedelta(days=90)
                interval = 'week'
            elif period == 'year':
                start_date = end_date - timedelta(days=365)
                interval = 'month'
            else:
                start_date = end_date - timedelta(days=30)
                interval = 'day'
            
            # 获取现金流数据
            cash_flow_data = self._get_cash_flow_by_interval(user_id, start_date, end_date, interval)
            
            # 计算现金流指标
            cash_flows = [item['net_flow'] for item in cash_flow_data]
            if cash_flows:
                # 现金流稳定性（标准差）
                cash_flow_volatility = float(np.std(cash_flows))
                
                # 现金流趋势（线性回归斜率）
                x = np.arange(len(cash_flows))
                if len(cash_flows) > 1:
                    slope, _ = np.polyfit(x, cash_flows, 1)
                    cash_flow_trend = float(slope)
                else:
                    cash_flow_trend = 0.0
                
                # 正现金流比例
                positive_periods = len([cf for cf in cash_flows if cf > 0])
                positive_ratio = positive_periods / len(cash_flows)
            else:
                cash_flow_volatility = 0.0
                cash_flow_trend = 0.0
                positive_ratio = 0.0
            
            # 现金流分类分析
            inflow_analysis = self._analyze_cash_inflows(user_id, start_date, end_date)
            outflow_analysis = self._analyze_cash_outflows(user_id, start_date, end_date)
            
            return {
                'period': period,
                'interval': interval,
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'cash_flow_data': cash_flow_data,
                'metrics': {
                    'volatility': cash_flow_volatility,
                    'trend': cash_flow_trend,
                    'positive_ratio': positive_ratio,
                    'total_periods': len(cash_flows)
                },
                'inflow_analysis': inflow_analysis,
                'outflow_analysis': outflow_analysis
            }
            
        except Exception as e:
            logger.error(f"现金流分析失败: {e}")
            return {}  
  
    # ==================== 报表生成 ====================
    
    def generate_transaction_report(self, user_id: int, report_type: str = 'summary',
                                  start_date: datetime = None, end_date: datetime = None) -> Dict[str, Any]:
        """生成交易报表"""
        try:
            if not start_date:
                start_date = datetime.now() - timedelta(days=30)
            if not end_date:
                end_date = datetime.now()
            
            # 获取基础数据
            transactions = self.get_transactions_by_user(
                user_id=user_id,
                filters={
                    'start_date': start_date,
                    'end_date': end_date,
                    'status': [TransactionStatus.COMPLETED]
                },
                limit=10000  # 报表需要完整数据
            )
            
            if report_type == 'summary':
                return self._generate_summary_report(transactions, start_date, end_date)
            elif report_type == 'detailed':
                return self._generate_detailed_report(transactions, start_date, end_date)
            elif report_type == 'tax':
                return self._generate_tax_report(transactions, start_date, end_date)
            elif report_type == 'audit':
                return self._generate_audit_report(transactions, start_date, end_date)
            else:
                return self._generate_summary_report(transactions, start_date, end_date)
                
        except Exception as e:
            logger.error(f"生成交易报表失败: {e}")
            return {}
    
    def export_transactions(self, user_id: int, export_format: str = 'csv',
                          filters: Dict[str, Any] = None) -> Dict[str, Any]:
        """导出交易流水"""
        try:
            # 获取交易数据
            search_params = {'user_id': user_id, 'limit': 50000}  # 导出限制
            if filters:
                search_params.update(filters)
            
            result = self.search_transactions(search_params)
            transactions = result['transactions']
            
            if not transactions:
                return {'success': False, 'error': '没有数据可导出'}
            
            # 准备导出数据
            export_data = []
            for transaction in transactions:
                export_data.append({
                    'transaction_id': transaction.transaction_id,
                    'account_id': transaction.account_id,
                    'transaction_type': transaction.transaction_type.value,
                    'status': transaction.status.value,
                    'amount': float(transaction.amount),
                    'currency': transaction.currency,
                    'balance_before': float(transaction.balance_before or 0),
                    'balance_after': float(transaction.balance_after or 0),
                    'symbol': transaction.symbol or '',
                    'description': transaction.description or '',
                    'fee_amount': float(transaction.fee_amount),
                    'tax_amount': float(transaction.tax_amount),
                    'transaction_time': transaction.transaction_time.isoformat() if transaction.transaction_time else '',
                    'reference_id': transaction.reference_id or ''
                })
            
            # 生成文件内容
            if export_format.lower() == 'csv':
                file_content = self._generate_csv_content(export_data)
                file_name = f"transactions_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                content_type = 'text/csv'
            elif export_format.lower() == 'excel':
                file_content = self._generate_excel_content(export_data)
                file_name = f"transactions_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif export_format.lower() == 'json':
                import json
                file_content = json.dumps(export_data, indent=2, ensure_ascii=False)
                file_name = f"transactions_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                content_type = 'application/json'
            else:
                return {'success': False, 'error': '不支持的导出格式'}
            
            return {
                'success': True,
                'file_name': file_name,
                'file_content': file_content,
                'content_type': content_type,
                'record_count': len(export_data)
            }
            
        except Exception as e:
            logger.error(f"导出交易流水失败: {e}")
            return {'success': False, 'error': str(e)}
    
    # ==================== 审计功能 ====================
    
    def audit_transactions(self, user_id: int, audit_type: str = 'consistency') -> Dict[str, Any]:
        """交易流水审计"""
        try:
            if audit_type == 'consistency':
                return self._audit_transaction_consistency(user_id)
            elif audit_type == 'completeness':
                return self._audit_transaction_completeness(user_id)
            elif audit_type == 'accuracy':
                return self._audit_transaction_accuracy(user_id)
            elif audit_type == 'timeline':
                return self._audit_transaction_timeline(user_id)
            else:
                return self._audit_transaction_consistency(user_id)
                
        except Exception as e:
            logger.error(f"交易流水审计失败: {e}")
            return {'error': str(e)}
    
    def get_suspicious_transactions(self, user_id: int, days: int = 30) -> List[Dict[str, Any]]:
        """获取可疑交易"""
        try:
            start_date = datetime.now() - timedelta(days=days)
            suspicious_transactions = []
            
            # 获取用户交易数据
            transactions = self.get_transactions_by_user(
                user_id=user_id,
                filters={'start_date': start_date},
                limit=10000
            )
            
            if not transactions:
                return suspicious_transactions
            
            # 计算统计指标
            amounts = [float(abs(t.amount)) for t in transactions]
            mean_amount = np.mean(amounts)
            std_amount = np.std(amounts)
            
            for transaction in transactions:
                suspicion_reasons = []
                
                # 异常金额检测
                amount = float(abs(transaction.amount))
                if amount > mean_amount + 3 * std_amount:
                    suspicion_reasons.append(f"交易金额异常大: {amount:.2f}")
                
                # 异常时间检测
                if transaction.transaction_time:
                    hour = transaction.transaction_time.hour
                    if hour < 6 or hour > 22:  # 非正常交易时间
                        suspicion_reasons.append(f"非正常交易时间: {hour}:00")
                
                # 重复交易检测
                duplicate_count = len([
                    t for t in transactions 
                    if t.amount == transaction.amount 
                    and t.transaction_time 
                    and transaction.transaction_time
                    and abs((t.transaction_time - transaction.transaction_time).total_seconds()) < 60
                    and t.id != transaction.id
                ])
                if duplicate_count > 0:
                    suspicion_reasons.append(f"疑似重复交易: {duplicate_count + 1}笔相同金额")
                
                # 状态异常检测
                if transaction.status in [TransactionStatus.FAILED, TransactionStatus.CANCELLED]:
                    if transaction.amount != 0:  # 失败或取消的交易不应该有金额变动
                        suspicion_reasons.append("失败/取消交易存在金额变动")
                
                if suspicion_reasons:
                    suspicious_transactions.append({
                        'transaction_id': transaction.transaction_id,
                        'transaction_time': transaction.transaction_time.isoformat() if transaction.transaction_time else None,
                        'amount': float(transaction.amount),
                        'transaction_type': transaction.transaction_type.value,
                        'status': transaction.status.value,
                        'description': transaction.description,
                        'suspicion_reasons': suspicion_reasons,
                        'risk_level': len(suspicion_reasons)  # 风险等级基于可疑原因数量
                    })
            
            # 按风险等级排序
            suspicious_transactions.sort(key=lambda x: x['risk_level'], reverse=True)
            
            return suspicious_transactions
            
        except Exception as e:
            logger.error(f"获取可疑交易失败: {e}")
            return []
    
    # ==================== 私有辅助方法 ====================
    
    def _generate_transaction_id(self) -> str:
        """生成交易ID"""
        import uuid
        return f"TXN_{datetime.now().strftime('%Y%m%d')}_{str(uuid.uuid4())[:8].upper()}"
    
    def _apply_transaction_filters(self, query, filters: Dict[str, Any]):
        """应用交易筛选条件"""
        if 'transaction_types' in filters:
            query = query.filter(Transaction.transaction_type.in_(filters['transaction_types']))
        
        if 'status' in filters:
            if isinstance(filters['status'], list):
                query = query.filter(Transaction.status.in_(filters['status']))
            else:
                query = query.filter(Transaction.status == filters['status'])
        
        if 'start_date' in filters:
            query = query.filter(Transaction.transaction_time >= filters['start_date'])
        
        if 'end_date' in filters:
            query = query.filter(Transaction.transaction_time <= filters['end_date'])
        
        if 'min_amount' in filters:
            query = query.filter(Transaction.amount >= filters['min_amount'])
        
        if 'max_amount' in filters:
            query = query.filter(Transaction.amount <= filters['max_amount'])
        
        if 'symbol' in filters:
            query = query.filter(Transaction.symbol == filters['symbol'])
        
        return query
    
    def _get_daily_transaction_stats(self, user_id: int, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """获取每日交易统计"""
        try:
            # 使用SQL按日期分组统计
            daily_stats = self.db.query(
                func.date(Transaction.transaction_time).label('date'),
                func.count(Transaction.id).label('count'),
                func.sum(Transaction.amount).label('total_amount'),
                func.sum(func.abs(Transaction.amount)).label('total_volume')
            ).join(Account).filter(
                Account.user_id == user_id,
                Transaction.transaction_time >= start_date,
                Transaction.transaction_time <= end_date,
                Transaction.status == TransactionStatus.COMPLETED
            ).group_by(func.date(Transaction.transaction_time)).all()
            
            return [
                {
                    'date': stat.date.isoformat() if stat.date else None,
                    'count': stat.count,
                    'total_amount': float(stat.total_amount or 0),
                    'total_volume': float(stat.total_volume or 0)
                }
                for stat in daily_stats
            ]
            
        except Exception as e:
            logger.error(f"获取每日交易统计失败: {e}")
            return []
    
    def _get_transaction_type_breakdown(self, user_id: int, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """获取交易类型分解"""
        try:
            type_stats = self.db.query(
                Transaction.transaction_type,
                func.count(Transaction.id).label('count'),
                func.sum(Transaction.amount).label('total_amount'),
                func.sum(Transaction.fee_amount).label('total_fees')
            ).join(Account).filter(
                Account.user_id == user_id,
                Transaction.transaction_time >= start_date,
                Transaction.transaction_time <= end_date,
                Transaction.status == TransactionStatus.COMPLETED
            ).group_by(Transaction.transaction_type).all()
            
            return [
                {
                    'type': stat.transaction_type.value,
                    'count': stat.count,
                    'total_amount': float(stat.total_amount or 0),
                    'total_fees': float(stat.total_fees or 0)
                }
                for stat in type_stats
            ]
            
        except Exception as e:
            logger.error(f"获取交易类型分解失败: {e}")
            return []
    
    def _get_cash_flow_by_interval(self, user_id: int, start_date: datetime, end_date: datetime, interval: str) -> List[Dict[str, Any]]:
        """按时间间隔获取现金流"""
        try:
            if interval == 'day':
                date_format = func.date(Transaction.transaction_time)
            elif interval == 'week':
                date_format = func.date_trunc('week', Transaction.transaction_time)
            elif interval == 'month':
                date_format = func.date_trunc('month', Transaction.transaction_time)
            else:
                date_format = func.date(Transaction.transaction_time)
            
            cash_flow_stats = self.db.query(
                date_format.label('period'),
                func.sum(func.case([(Transaction.amount > 0, Transaction.amount)], else_=0)).label('inflow'),
                func.sum(func.case([(Transaction.amount < 0, func.abs(Transaction.amount))], else_=0)).label('outflow'),
                func.sum(Transaction.amount).label('net_flow')
            ).join(Account).filter(
                Account.user_id == user_id,
                Transaction.transaction_time >= start_date,
                Transaction.transaction_time <= end_date,
                Transaction.status == TransactionStatus.COMPLETED
            ).group_by(date_format).order_by(date_format).all()
            
            return [
                {
                    'period': stat.period.isoformat() if stat.period else None,
                    'inflow': float(stat.inflow or 0),
                    'outflow': float(stat.outflow or 0),
                    'net_flow': float(stat.net_flow or 0)
                }
                for stat in cash_flow_stats
            ]
            
        except Exception as e:
            logger.error(f"按时间间隔获取现金流失败: {e}")
            return []
    
    def _analyze_cash_inflows(self, user_id: int, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """分析现金流入"""
        try:
            inflow_stats = self.db.query(
                Transaction.transaction_type,
                func.count(Transaction.id).label('count'),
                func.sum(Transaction.amount).label('total_amount'),
                func.avg(Transaction.amount).label('avg_amount')
            ).join(Account).filter(
                Account.user_id == user_id,
                Transaction.transaction_time >= start_date,
                Transaction.transaction_time <= end_date,
                Transaction.amount > 0,
                Transaction.status == TransactionStatus.COMPLETED
            ).group_by(Transaction.transaction_type).all()
            
            total_inflow = sum(float(stat.total_amount or 0) for stat in inflow_stats)
            
            return {
                'total_inflow': total_inflow,
                'by_type': [
                    {
                        'type': stat.transaction_type.value,
                        'count': stat.count,
                        'total_amount': float(stat.total_amount or 0),
                        'avg_amount': float(stat.avg_amount or 0),
                        'percentage': float(stat.total_amount or 0) / total_inflow * 100 if total_inflow > 0 else 0
                    }
                    for stat in inflow_stats
                ]
            }
            
        except Exception as e:
            logger.error(f"分析现金流入失败: {e}")
            return {}
    
    def _analyze_cash_outflows(self, user_id: int, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """分析现金流出"""
        try:
            outflow_stats = self.db.query(
                Transaction.transaction_type,
                func.count(Transaction.id).label('count'),
                func.sum(func.abs(Transaction.amount)).label('total_amount'),
                func.avg(func.abs(Transaction.amount)).label('avg_amount')
            ).join(Account).filter(
                Account.user_id == user_id,
                Transaction.transaction_time >= start_date,
                Transaction.transaction_time <= end_date,
                Transaction.amount < 0,
                Transaction.status == TransactionStatus.COMPLETED
            ).group_by(Transaction.transaction_type).all()
            
            total_outflow = sum(float(stat.total_amount or 0) for stat in outflow_stats)
            
            return {
                'total_outflow': total_outflow,
                'by_type': [
                    {
                        'type': stat.transaction_type.value,
                        'count': stat.count,
                        'total_amount': float(stat.total_amount or 0),
                        'avg_amount': float(stat.avg_amount or 0),
                        'percentage': float(stat.total_amount or 0) / total_outflow * 100 if total_outflow > 0 else 0
                    }
                    for stat in outflow_stats
                ]
            }
            
        except Exception as e:
            logger.error(f"分析现金流出失败: {e}")
            return {}
    
    def _generate_summary_report(self, transactions: List[Transaction], start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """生成汇总报表"""
        try:
            if not transactions:
                return {'error': '没有交易数据'}
            
            total_count = len(transactions)
            total_income = sum(float(t.amount) for t in transactions if t.amount > 0)
            total_expense = sum(float(abs(t.amount)) for t in transactions if t.amount < 0)
            total_fees = sum(float(t.fee_amount) for t in transactions)
            
            # 按类型统计
            type_summary = {}
            for transaction in transactions:
                tx_type = transaction.transaction_type.value
                if tx_type not in type_summary:
                    type_summary[tx_type] = {'count': 0, 'amount': 0}
                type_summary[tx_type]['count'] += 1
                type_summary[tx_type]['amount'] += float(transaction.amount)
            
            return {
                'report_type': 'summary',
                'period': {
                    'start_date': start_date.isoformat(),
                    'end_date': end_date.isoformat()
                },
                'summary': {
                    'total_transactions': total_count,
                    'total_income': total_income,
                    'total_expense': total_expense,
                    'net_amount': total_income - total_expense,
                    'total_fees': total_fees
                },
                'by_type': type_summary,
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"生成汇总报表失败: {e}")
            return {'error': str(e)}
    
    def _generate_detailed_report(self, transactions: List[Transaction], start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """生成详细报表"""
        try:
            detailed_data = []
            for transaction in transactions:
                detailed_data.append({
                    'transaction_id': transaction.transaction_id,
                    'date': transaction.transaction_time.isoformat() if transaction.transaction_time else None,
                    'type': transaction.transaction_type.value,
                    'amount': float(transaction.amount),
                    'currency': transaction.currency,
                    'description': transaction.description,
                    'status': transaction.status.value,
                    'fee_amount': float(transaction.fee_amount),
                    'balance_after': float(transaction.balance_after or 0)
                })
            
            return {
                'report_type': 'detailed',
                'period': {
                    'start_date': start_date.isoformat(),
                    'end_date': end_date.isoformat()
                },
                'transactions': detailed_data,
                'total_count': len(detailed_data),
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"生成详细报表失败: {e}")
            return {'error': str(e)}
    
    def _generate_tax_report(self, transactions: List[Transaction], start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """生成税务报表"""
        try:
            tax_data = []
            total_tax = 0
            
            for transaction in transactions:
                if transaction.tax_amount and transaction.tax_amount > 0:
                    tax_data.append({
                        'transaction_id': transaction.transaction_id,
                        'date': transaction.transaction_time.isoformat() if transaction.transaction_time else None,
                        'type': transaction.transaction_type.value,
                        'amount': float(transaction.amount),
                        'tax_amount': float(transaction.tax_amount),
                        'description': transaction.description
                    })
                    total_tax += float(transaction.tax_amount)
            
            return {
                'report_type': 'tax',
                'period': {
                    'start_date': start_date.isoformat(),
                    'end_date': end_date.isoformat()
                },
                'tax_summary': {
                    'total_tax_amount': total_tax,
                    'taxable_transactions': len(tax_data)
                },
                'tax_details': tax_data,
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"生成税务报表失败: {e}")
            return {'error': str(e)}
    
    def _generate_audit_report(self, transactions: List[Transaction], start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """生成审计报表"""
        try:
            audit_issues = []
            
            # 检查余额一致性
            for i, transaction in enumerate(transactions[:-1]):
                next_transaction = transactions[i + 1]
                if (transaction.balance_after and next_transaction.balance_before and 
                    transaction.balance_after != next_transaction.balance_before):
                    audit_issues.append({
                        'type': 'balance_inconsistency',
                        'transaction_id': transaction.transaction_id,
                        'issue': f"余额不一致: {transaction.balance_after} != {next_transaction.balance_before}"
                    })
            
            # 检查状态异常
            for transaction in transactions:
                if transaction.status == TransactionStatus.FAILED and transaction.amount != 0:
                    audit_issues.append({
                        'type': 'status_amount_mismatch',
                        'transaction_id': transaction.transaction_id,
                        'issue': "失败交易存在金额变动"
                    })
            
            return {
                'report_type': 'audit',
                'period': {
                    'start_date': start_date.isoformat(),
                    'end_date': end_date.isoformat()
                },
                'audit_summary': {
                    'total_transactions': len(transactions),
                    'issues_found': len(audit_issues),
                    'audit_status': 'PASSED' if len(audit_issues) == 0 else 'FAILED'
                },
                'issues': audit_issues,
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"生成审计报表失败: {e}")
            return {'error': str(e)}
    
    def _generate_csv_content(self, data: List[Dict[str, Any]]) -> str:
        """生成CSV内容"""
        try:
            import csv
            import io
            
            if not data:
                return ""
            
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"生成CSV内容失败: {e}")
            return ""
    
    def _generate_excel_content(self, data: List[Dict[str, Any]]) -> bytes:
        """生成Excel内容"""
        try:
            import io
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "交易流水"
            
            if data:
                # 写入表头
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # 写入数据
                for row, item in enumerate(data, 2):
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=item[header])
            
            # 保存到字节流
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"生成Excel内容失败: {e}")
            return b""
    
    def _audit_transaction_consistency(self, user_id: int) -> Dict[str, Any]:
        """审计交易一致性"""
        try:
            issues = []
            
            # 获取用户所有账户的交易
            transactions = self.get_transactions_by_user(user_id, limit=10000)
            
            # 按账户分组检查余额一致性
            account_transactions = {}
            for transaction in transactions:
                account_id = transaction.account_id
                if account_id not in account_transactions:
                    account_transactions[account_id] = []
                account_transactions[account_id].append(transaction)
            
            for account_id, account_txns in account_transactions.items():
                # 按时间排序
                account_txns.sort(key=lambda x: x.transaction_time or datetime.min)
                
                # 检查余额连续性
                for i in range(len(account_txns) - 1):
                    current = account_txns[i]
                    next_txn = account_txns[i + 1]
                    
                    if (current.balance_after and next_txn.balance_before and 
                        current.balance_after != next_txn.balance_before):
                        issues.append({
                            'type': 'balance_inconsistency',
                            'account_id': account_id,
                            'transaction_id': current.transaction_id,
                            'next_transaction_id': next_txn.transaction_id,
                            'description': f"余额不连续: {current.balance_after} -> {next_txn.balance_before}"
                        })
            
            return {
                'audit_type': 'consistency',
                'total_issues': len(issues),
                'issues': issues,
                'status': 'PASSED' if len(issues) == 0 else 'FAILED',
                'audited_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"审计交易一致性失败: {e}")
            return {'error': str(e)}
    
    def _audit_transaction_completeness(self, user_id: int) -> Dict[str, Any]:
        """审计交易完整性"""
        try:
            issues = []
            
            # 检查必填字段
            transactions = self.db.query(Transaction).join(Account).filter(
                Account.user_id == user_id
            ).all()
            
            for transaction in transactions:
                if not transaction.transaction_id:
                    issues.append({
                        'type': 'missing_transaction_id',
                        'transaction_id': transaction.id,
                        'description': '缺少交易ID'
                    })
                
                if not transaction.transaction_time:
                    issues.append({
                        'type': 'missing_transaction_time',
                        'transaction_id': transaction.transaction_id,
                        'description': '缺少交易时间'
                    })
                
                if transaction.amount is None:
                    issues.append({
                        'type': 'missing_amount',
                        'transaction_id': transaction.transaction_id,
                        'description': '缺少交易金额'
                    })
            
            return {
                'audit_type': 'completeness',
                'total_issues': len(issues),
                'issues': issues,
                'status': 'PASSED' if len(issues) == 0 else 'FAILED',
                'audited_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"审计交易完整性失败: {e}")
            return {'error': str(e)}
    
    def _audit_transaction_accuracy(self, user_id: int) -> Dict[str, Any]:
        """审计交易准确性"""
        try:
            issues = []
            
            transactions = self.get_transactions_by_user(user_id, limit=10000)
            
            for transaction in transactions:
                # 检查金额计算准确性
                if (transaction.balance_before is not None and 
                    transaction.balance_after is not None and
                    transaction.amount is not None):
                    
                    expected_balance = transaction.balance_before + transaction.amount
                    if abs(expected_balance - transaction.balance_after) > Decimal('0.01'):
                        issues.append({
                            'type': 'balance_calculation_error',
                            'transaction_id': transaction.transaction_id,
                            'description': f"余额计算错误: {transaction.balance_before} + {transaction.amount} != {transaction.balance_after}"
                        })
                
                # 检查手续费合理性
                if transaction.fee_amount and transaction.fee_amount < 0:
                    issues.append({
                        'type': 'negative_fee',
                        'transaction_id': transaction.transaction_id,
                        'description': f"手续费为负数: {transaction.fee_amount}"
                    })
            
            return {
                'audit_type': 'accuracy',
                'total_issues': len(issues),
                'issues': issues,
                'status': 'PASSED' if len(issues) == 0 else 'FAILED',
                'audited_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"审计交易准确性失败: {e}")
            return {'error': str(e)}
    
    def _audit_transaction_timeline(self, user_id: int) -> Dict[str, Any]:
        """审计交易时间线"""
        try:
            issues = []
            
            transactions = self.get_transactions_by_user(user_id, limit=10000)
            
            # 检查时间顺序
            transactions.sort(key=lambda x: x.transaction_time or datetime.min)
            
            for i in range(len(transactions) - 1):
                current = transactions[i]
                next_txn = transactions[i + 1]
                
                if (current.transaction_time and next_txn.transaction_time and
                    current.transaction_time > next_txn.transaction_time):
                    issues.append({
                        'type': 'timeline_disorder',
                        'transaction_id': current.transaction_id,
                        'next_transaction_id': next_txn.transaction_id,
                        'description': f"交易时间顺序错误: {current.transaction_time} > {next_txn.transaction_time}"
                    })
            
            # 检查未来时间
            now = datetime.now()
            for transaction in transactions:
                if transaction.transaction_time and transaction.transaction_time > now:
                    issues.append({
                        'type': 'future_transaction',
                        'transaction_id': transaction.transaction_id,
                        'description': f"交易时间在未来: {transaction.transaction_time}"
                    })
            
            return {
                'audit_type': 'timeline',
                'total_issues': len(issues),
                'issues': issues,
                'status': 'PASSED' if len(issues) == 0 else 'FAILED',
                'audited_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"审计交易时间线失败: {e}")
            return {'error': str(e)}